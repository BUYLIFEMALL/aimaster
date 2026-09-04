"""Phase 4 E2E 통합 테스트: Spider 대규모 수집 + ProgressTracker + 중간 저장"""
import json
import os
import shutil
import sys
from typing import Any, Dict

sys.path.insert(0, os.path.dirname(__file__))
from utils import setup_logger, sanitize_filename
from export_excel import export_to_excel
from progress import ProgressTracker
from datetime import datetime


def test_e2e_spider_collection(tmp_path):
    """Spider로 books.toscrape.com 전체를 수집 + ProgressTracker 통합.

    books.toscrape.com은 1000건(50페이지 x 20건).
    테스트에서는 5페이지(100건)로 제한하여 검증.
    """
    from scrapling.spiders import Spider, Request, Response

    logger = setup_logger("e2e_phase4_spider")
    # 산출물은 tmp_path 아래로. 실제 output/ 에 쓰면 도메인별 폴더 규약을 어기고
    # pytest 를 돌릴 때마다 output/ 루트에 파편이 쌓인다.
    output_dir = str(tmp_path)
    crawl_dir = str(tmp_path / "crawl_data")   # repo 루트에 crawl_data_test 를 만들지 않는다

    # Clean up previous test crawl data
    if os.path.exists(crawl_dir):
        shutil.rmtree(crawl_dir)

    # === ProgressTracker 설정 ===
    progress_path = os.path.join(output_dir, "progress_test.json")
    tracker = ProgressTracker(progress_path, total_estimate=100)

    logger.info("=== Phase 4 E2E: Spider Collection (books.toscrape.com) ===")

    item_count = 0
    max_pages = 5

    class BookSpider(Spider):
        name = "book_collector"
        start_urls = ["https://books.toscrape.com/catalogue/page-1.html"]
        concurrent_requests = 3

        async def parse(self, response: Response):
            nonlocal item_count

            for book in response.css("article.product_pod"):
                title = book.css("h3 a::attr(title)").get("").strip()
                price = book.css("p.price_color::text").get("").strip()
                availability = book.css("p.instock.availability::text").getall()
                availability = " ".join(a.strip() for a in availability).strip()
                rating_class = book.css("p.star-rating::attr(class)").get("")
                rating = rating_class.replace("star-rating ", "") if rating_class else ""

                yield {
                    "제목": title,
                    "가격": price,
                    "재고": availability,
                    "평점": rating,
                }

            # Pagination: next page
            next_href = response.css("li.next a::attr(href)").get()
            if next_href:
                # Extract page number from href
                page_match = next_href
                import re
                page_num_match = re.search(r'page-(\d+)', next_href)
                if page_num_match:
                    page_num = int(page_num_match.group(1))
                    if page_num <= max_pages:
                        yield response.follow(next_href)

        async def on_scraped_item(self, item: Dict[str, Any]) -> Dict[str, Any] | None:
            nonlocal item_count
            item_count += 1
            if item_count % 20 == 0:
                tracker.update(collected=item_count, current_page=item_count // 20)
            return item

        async def on_error(self, request: Request, error: Exception) -> None:
            tracker.add_error(f"Request error: {error}")
            logger.warning(f"Spider error: {error}")

        async def on_close(self) -> None:
            tracker.update(collected=item_count)
            tracker.complete()
            logger.info(f"Spider closed. Total items: {item_count}")

    # === Spider 실행 ===
    result = BookSpider(crawldir=crawl_dir).start()

    # === 검증 ===
    logger.info(f"Spider result: {len(result.items)} items, completed={result.completed}")

    # 1. 아이템 수 확인 (5페이지 x 20건 = 100건)
    assert len(result.items) >= 80, f"Expected >= 80 items, got {len(result.items)}"
    assert len(result.items) <= 100, f"Expected <= 100 items, got {len(result.items)}"
    logger.info(f"Item count: {len(result.items)} (expected ~100)")

    # 2. 데이터 품질 확인
    empty_titles = sum(1 for item in result.items if not item.get("제목"))
    assert empty_titles == 0, f"{empty_titles} empty titles"
    empty_prices = sum(1 for item in result.items if not item.get("가격"))
    assert empty_prices == 0, f"{empty_prices} empty prices"

    # 3. 첫/마지막 아이템 확인
    items_list = list(result.items)
    logger.info(f"First item: {items_list[0]}")
    logger.info(f"Last item: {items_list[-1]}")

    # 4. ProgressTracker 상태 확인
    progress = tracker.read()
    assert progress["status"] == "completed", f"Progress status: {progress['status']}"
    assert progress["collected"] > 0, f"Progress collected: {progress['collected']}"
    logger.info(f"Progress: {progress['collected']}/{progress['total_estimate']} items, status={progress['status']}")

    # 5. JSON 출력
    raw_path = os.path.join(output_dir, "raw_data_spider.json")
    result.items.to_json(raw_path, indent=True)
    assert os.path.exists(raw_path)
    with open(raw_path, "r", encoding="utf-8") as f:
        saved_data = json.load(f)
    assert len(saved_data) == len(result.items)
    logger.info(f"Raw data saved: {raw_path}")

    # 6. 엑셀 출력
    domain = sanitize_filename("books_toscrape_com_spider")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f"crawl_{domain}_{timestamp}.xlsx")
    export_result = export_to_excel(saved_data, excel_path)
    assert export_result is not None
    assert os.path.exists(excel_path)

    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    assert ws.cell(1, 1).value == "제목"
    assert ws.max_row >= 81  # header + 80+ rows
    logger.info(f"Excel saved: {excel_path} ({ws.max_row - 1} rows)")

    # 7. crawldir: Spider는 완료 시 checkpoint를 자동 정리함
    # 완료된 경우 crawldir이 없을 수 있음 (정상 동작)
    logger.info(f"Crawl dir exists after completion: {os.path.exists(crawl_dir)} (cleaned up is normal)")

    # Cleanup
    if os.path.exists(crawl_dir):
        shutil.rmtree(crawl_dir)
    if os.path.exists(progress_path):
        os.remove(progress_path)

    logger.info("=== Spider Collection Test PASSED ===")


def test_e2e_session_checkpoint(tmp_path):
    """Session 기반 중간 저장 + Resume 패턴 검증."""
    from scrapling.fetchers import Fetcher

    logger = setup_logger("e2e_phase4_checkpoint")
    # 산출물은 tmp_path 아래로. 실제 output/ 에 쓰면 도메인별 폴더 규약을 어기고
    # pytest 를 돌릴 때마다 output/ 루트에 파편이 쌓인다.
    output_dir = str(tmp_path)

    logger.info("=== Phase 4 E2E: Session Checkpoint + Resume ===")

    checkpoint_path = os.path.join(output_dir, "raw_data_checkpoint_test.json")

    # Clean previous test data
    if os.path.exists(checkpoint_path):
        os.remove(checkpoint_path)

    # === Phase A: 수집 2페이지 + 중간 저장 ===
    from utils import RateLimiter
    fetcher = Fetcher()
    limiter = RateLimiter(delay=0.5)
    results = []
    items_per_page = 20

    for page_num in range(1, 3):  # 2페이지만
        limiter.wait()
        url = f"https://books.toscrape.com/catalogue/page-{page_num}.html"
        page = fetcher.get(url)
        assert page.status == 200

        for book in page.css("article.product_pod"):
            results.append({
                "제목": book.css("h3 a::attr(title)").get("").strip(),
                "가격": book.css("p.price_color::text").get("").strip(),
            })

        # 중간 저장 (100건마다이지만 테스트에서는 매 페이지)
        with open(checkpoint_path, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)

    assert len(results) == 40, f"Phase A: expected 40, got {len(results)}"
    logger.info(f"Phase A: collected {len(results)} items, checkpoint saved")

    # === Phase B: Resume — 기존 데이터 로드 + 이어서 수집 ===
    if os.path.exists(checkpoint_path):
        with open(checkpoint_path, "r", encoding="utf-8") as f:
            results = json.load(f)
        start_page = (len(results) // items_per_page) + 1
        logger.info(f"Resuming from page {start_page} ({len(results)} existing items)")
    else:
        results = []
        start_page = 1

    assert start_page == 3, f"Should resume from page 3, got {start_page}"

    for page_num in range(start_page, start_page + 2):  # 2페이지 더
        limiter.wait()
        url = f"https://books.toscrape.com/catalogue/page-{page_num}.html"
        page = fetcher.get(url)
        assert page.status == 200

        for book in page.css("article.product_pod"):
            results.append({
                "제목": book.css("h3 a::attr(title)").get("").strip(),
                "가격": book.css("p.price_color::text").get("").strip(),
            })

        with open(checkpoint_path, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)

    assert len(results) == 80, f"Phase B: expected 80, got {len(results)}"
    logger.info(f"Phase B: total {len(results)} items after resume")

    # 중복 확인
    titles = [r["제목"] for r in results]
    unique_titles = set(titles)
    assert len(unique_titles) == len(titles), f"Duplicates found: {len(titles) - len(unique_titles)}"
    logger.info(f"No duplicates: {len(unique_titles)} unique titles")

    # Cleanup
    if os.path.exists(checkpoint_path):
        os.remove(checkpoint_path)

    logger.info("=== Session Checkpoint + Resume Test PASSED ===")


if __name__ == "__main__":
    test_e2e_session_checkpoint()
    print()
    test_e2e_spider_collection()
    print("\n=== ALL PHASE 4 E2E TESTS PASSED ===")
