"""Phase 1 E2E 통합 테스트: books.toscrape.com 정찰→수집→엑셀 출력"""
import json
import os
import sys
import time

sys.path.insert(0, os.path.dirname(__file__))
from utils import RateLimiter, setup_logger, sanitize_filename
from export_excel import export_to_excel
from scrapling.fetchers import Fetcher
from datetime import datetime
from urllib.parse import urljoin


def test_e2e_books_toscrape(tmp_path):
    """books.toscrape.com에서 책 제목, 가격, 평점을 수집하여 엑셀로 출력."""
    logger = setup_logger("e2e_test")
    # 산출물은 tmp_path 아래로. 실제 output/ 에 쓰면 도메인별 폴더 규약을 어기고
    # pytest 를 돌릴 때마다 output/ 루트에 파편이 쌓인다.
    output_dir = str(tmp_path)

    # === Step 1: 수집 ===
    logger.info("=== E2E Test Start: books.toscrape.com ===")
    limiter = RateLimiter(delay=1.0)
    results = []
    fetcher = Fetcher()
    base_url = "http://books.toscrape.com/"
    current_url = base_url
    page_num = 1
    max_pages = 3  # E2E 테스트에서는 3페이지만 수집 (속도)

    while page_num <= max_pages:
        limiter.wait()
        logger.info(f"Fetching page {page_num}: {current_url}")

        page = fetcher.get(current_url)
        if page.status != 200:
            logger.warning(f"Status {page.status}, stopping")
            break

        articles = page.css("article.product_pod")
        if not articles:
            logger.warning("No articles found, stopping")
            break

        for article in articles:
            title = article.css("h3 a::attr(title)").get("")
            price = article.css("p.price_color::text").get("")
            # Rating from class: "star-rating Three" -> "Three"
            rating_el = article.css("p.star-rating")
            rating = "Unknown"
            if rating_el:
                classes = rating_el[0].attrib.get("class", "")
                parts = classes.split()
                if len(parts) > 1:
                    rating = parts[1]

            results.append({
                "제목": title.strip(),
                "가격": price.strip(),
                "평점": rating,
            })

        logger.info(f"Page {page_num}: {len(articles)} items, total {len(results)}")

        # Pagination
        next_link = page.css("li.next a::attr(href)").get()
        if not next_link:
            break
        current_url = urljoin(current_url, next_link)
        page_num += 1

    # === Step 2: 원시 데이터 저장 ===
    raw_path = os.path.join(output_dir, "raw_data.json")
    with open(raw_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    logger.info(f"Raw data saved: {raw_path} ({len(results)} items)")

    # === Step 3: 데이터 검증 ===
    assert len(results) > 0, "No data collected"
    assert len(results) == 60, f"Expected 60 items (3 pages x 20), got {len(results)}"

    # 필드 유효성 체크
    empty_titles = sum(1 for r in results if not r["제목"])
    empty_prices = sum(1 for r in results if not r["가격"])
    empty_ratings = sum(1 for r in results if r["평점"] == "Unknown")
    logger.info(f"Empty titles: {empty_titles}, prices: {empty_prices}, ratings: {empty_ratings}")
    assert empty_titles == 0, f"{empty_titles} empty titles"
    assert empty_prices == 0, f"{empty_prices} empty prices"

    # 샘플 확인
    logger.info(f"First item: {results[0]}")
    logger.info(f"Last item: {results[-1]}")

    # 중복 체크
    titles = [r["제목"] for r in results]
    unique_titles = set(titles)
    dupes = len(titles) - len(unique_titles)
    logger.info(f"Duplicates: {dupes}")

    # === Step 4: 엑셀 출력 ===
    domain = sanitize_filename("books_toscrape_com")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f"crawl_{domain}_{timestamp}.xlsx")
    result = export_to_excel(results, excel_path)
    assert result is not None, "Excel export returned None"
    assert os.path.exists(excel_path), f"Excel file not found: {excel_path}"
    logger.info(f"Excel saved: {excel_path}")

    # === Step 5: 엑셀 내용 검증 ===
    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    assert ws.cell(1, 1).value == "제목", f"Header mismatch: {ws.cell(1, 1).value}"
    assert ws.cell(1, 2).value == "가격"
    assert ws.cell(1, 3).value == "평점"
    assert ws.max_row == len(results) + 1, f"Row count mismatch: {ws.max_row}"
    assert ws.cell(1, 1).font.bold is True, "Header not bold"
    assert ws.auto_filter.ref is not None, "Auto filter not set"
    logger.info(f"Excel verification passed: {ws.max_row - 1} data rows")

    logger.info("=== E2E Test PASSED ===")


if __name__ == "__main__":
    result = test_e2e_books_toscrape()
    print(f"\nResult: {result}")
