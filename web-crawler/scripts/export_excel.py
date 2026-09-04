"""scripts/export_excel.py — 수집 데이터를 엑셀 파일로 출력"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def export_to_excel(data: list[dict], filepath: str, sheet_name: str = "수집 데이터") -> str | None:
    """데이터 리스트를 엑셀 파일로 저장.

    Args:
        data: 딕셔너리 리스트. 각 딕셔너리의 키가 헤더가 됨.
        filepath: 저장할 .xlsx 파일 경로.
        sheet_name: 시트 이름 (기본값: "수집 데이터").

    Returns:
        저장된 파일 경로. 데이터가 없으면 None.
    """
    if not data:
        return None

    os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    headers = list(data[0].keys())

    # Header row (bold)
    bold_font = Font(bold=True)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font

    # Data rows
    for row_idx, item in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=item.get(header))

    # Auto column width
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in range(2, len(data) + 2):
            val = ws.cell(row=row, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 50)

    # Auto filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"

    wb.save(filepath)
    return filepath
