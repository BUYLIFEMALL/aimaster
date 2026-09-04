"""Phase 3 E2E 통합 테스트: DynamicFetcher + StealthyFetcher + 에스컬레이션"""
import json
import os
import sys

sys.path.insert(0, os.path.dirname(__file__))
from utils import RateLimiter, setup_logger, sanitize_filename
from export_excel import export_to_excel
from datetime import datetime


def test_e2e_dynamic_fetcher(tmp_path):
    """DynamicFetcher로 JS 렌더링이 필요한 사이트를 수집하여 엑셀로 출력.

    quotes.toscrape.com/js/ 는 JavaScript로 데이터를 렌더링하는 버전.
    일반 Fetcher로는 빈 페이지가 반환되고, DynamicFetcher로만 수집 가능.
    """
    from scrapling.fetchers import Fetcher, DynamicFetcher

    logger = setup_logger("e2e_phase3_dynamic")
    # 산출물은 tmp_path 아래로. 실제 output/ 에 쓰면 도메인별 폴더 규약을 어기고
    # pytest 를 돌릴 때마다 output/ 루트에 파편이 쌓인다.
    output_dir = str(tmp_path)

    # === Step 1: 일반 Fetcher로 실패 확인 ===
    logger.info("=== Phase 3 E2E: DynamicFetcher (quotes.toscrape.com/js/) ===")
    fetcher = Fetcher()
    page = fetcher.get("https://quotes.toscrape.com/js/")
    basic_quotes = page.css("div.quote")
    logger.info(f"Basic Fetcher found {len(basic_quotes)} quotes (expected 0 or few)")

    # === Step 2: DynamicFetcher로 수집 ===
    logger.info("Switching to DynamicFetcher for JS rendering...")
    dynamic = DynamicFetcher()
    limiter = RateLimiter(delay=1.5)
    results = []
    page_num = 1
    max_pages = 3

    while page_num <= max_pages:
        limiter.wait()
        url = f"https://quotes.toscrape.com/js/page/{page_num}/"
        logger.info(f"Rendering page {page_num}: {url}")

        page = dynamic.fetch(url, network_idle=True)
        if page.status != 200:
            logger.warning(f"Status {page.status}, stopping")
            break

        quotes = page.css("div.quote")
        if not quotes:
            logger.info("No more quotes found, stopping")
            break

        for quote in quotes:
            text = quote.css("span.text::text").get("").strip()
            author = quote.css("small.author::text").get("").strip()
            tags = [t.text.strip() for t in quote.css("a.tag")]
            results.append({
                "인용문": text,
                "저자": author,
                "태그": ", ".join(tags),
            })

        logger.info(f"Page {page_num}: {len(quotes)} quotes, total {len(results)}")
        page_num += 1

    # === Step 3: 데이터 검증 ===
    assert len(results) >= 20, f"Expected >= 20 quotes, got {len(results)}"
    empty_authors = sum(1 for r in results if not r["저자"])
    assert empty_authors == 0, f"{empty_authors} empty authors"

    # DynamicFetcher가 일반 Fetcher보다 더 많은 데이터를 가져왔는지 확인
    assert len(results) > len(basic_quotes), (
        f"DynamicFetcher ({len(results)}) should find more than Fetcher ({len(basic_quotes)})"
    )

    logger.info(f"First: {results[0]}")
    logger.info(f"Last: {results[-1]}")

    # === Step 4: 원시 데이터 저장 ===
    raw_path = os.path.join(output_dir, "raw_data_dynamic.json")
    with open(raw_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    logger.info(f"Raw data saved: {raw_path}")

    # === Step 5: 엑셀 출력 ===
    domain = sanitize_filename("quotes_toscrape_com_js")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f"crawl_{domain}_{timestamp}.xlsx")
    result = export_to_excel(results, excel_path)
    assert result is not None
    assert os.path.exists(excel_path)

    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    assert ws.cell(1, 1).value == "인용문"
    assert ws.max_row >= 21  # header + 20+ rows
    logger.info(f"Excel saved: {excel_path} ({ws.max_row - 1} rows)")

    logger.info("=== DynamicFetcher Test PASSED ===")


def test_e2e_escalation_chain():
    """에스컬레이션 체인: Fetcher 실패 → StealthyFetcher → DynamicFetcher 자동 전환.

    quotes.toscrape.com/js/ 는 JS 렌더링 필요 → Fetcher 부분 실패 → 에스컬레이션.
    """
    from scrapling.fetchers import DynamicFetcher

    from utils import plain_get, plain_session

    logger = setup_logger("e2e_phase3_escalation")

    def _session_tier(url):
        with plain_session() as session:
            return session.get(url)

    # 사다리 A 전용 — StealthyFetcher(5단)는 통지 게이트 뒤에 있으므로 자동 체인에 없다
    FETCHER_CHAIN = [
        ("plain_get", plain_get),
        ("plain_session", _session_tier),
        ("DynamicFetcher", lambda url: DynamicFetcher().fetch(url, network_idle=True)),
    ]

    def fetch_with_escalation(url: str):
        for name, fetch_fn in FETCHER_CHAIN:
            try:
                page = fetch_fn(url)
                if page.status == 200:
                    # JS 사이트: quote가 있어야 성공
                    quotes = page.css("div.quote")
                    if quotes:
                        logger.info(f"[{name}] Success - found {len(quotes)} quotes")
                        return page, name
                    else:
                        logger.warning(f"[{name}] Page loaded but no quotes found, escalating")
                        continue
                else:
                    logger.warning(f"[{name}] Status {page.status}, escalating")
                    continue
            except Exception as e:
                logger.warning(f"[{name}] Error: {e}, escalating")
                continue
        return None, None

    # === 에스컬레이션 테스트 ===
    logger.info("=== Phase 3 E2E: Fetcher Escalation Chain ===")
    url = "https://quotes.toscrape.com/js/"

    page, used_fetcher = fetch_with_escalation(url)

    assert page is not None, "All fetchers failed - escalation chain broken"
    assert used_fetcher is not None
    logger.info(f"Final fetcher used: {used_fetcher}")

    # DynamicFetcher 또는 StealthyFetcher가 사용되었어야 함
    # (기본 Fetcher로는 JS 사이트 수집 불가능할 수 있음)
    quotes = page.css("div.quote")
    assert len(quotes) > 0, f"No quotes found even after escalation"
    logger.info(f"Escalation result: {len(quotes)} quotes via {used_fetcher}")

    logger.info("=== Escalation Chain Test PASSED ===")


def test_e2e_stealthy_fetcher():
    """StealthyFetcher로 정상 사이트 수집 확인."""
    from scrapling.fetchers import StealthyFetcher

    logger = setup_logger("e2e_phase3_stealthy")

    logger.info("=== Phase 3 E2E: StealthyFetcher ===")
    fetcher = StealthyFetcher()
    page = fetcher.fetch("https://quotes.toscrape.com/", headless=True)

    assert page.status == 200, f"Status {page.status}"
    quotes = page.css("div.quote")
    assert len(quotes) > 0, f"No quotes found"

    first_text = quotes[0].css("span.text::text").get("").strip()
    first_author = quotes[0].css("small.author::text").get("").strip()
    logger.info(f"StealthyFetcher: {len(quotes)} quotes, first by {first_author}")
    assert first_author, "Empty author"

    logger.info("=== StealthyFetcher Test PASSED ===")


if __name__ == "__main__":
    test_e2e_stealthy_fetcher()
    print()
    test_e2e_escalation_chain()
    print()
    test_e2e_dynamic_fetcher()
    print("\n=== ALL PHASE 3 E2E TESTS PASSED ===")
