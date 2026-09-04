"""scripts/test_export_excel.py"""
import os
from openpyxl import load_workbook
from export_excel import export_to_excel


def test_export_creates_file(tmp_path):
    data = [
        {"상품명": "테스트 상품 A", "가격": 15000, "리뷰수": 42},
        {"상품명": "테스트 상품 B", "가격": 23000, "리뷰수": 17},
    ]
    filepath = str(tmp_path / "test_output.xlsx")
    result = export_to_excel(data, filepath)
    assert os.path.exists(result)


def test_export_correct_content(tmp_path):
    data = [
        {"상품명": "상품A", "가격": 10000},
        {"상품명": "상품B", "가격": 20000},
    ]
    filepath = str(tmp_path / "test.xlsx")
    export_to_excel(data, filepath)
    wb = load_workbook(filepath)
    ws = wb.active
    # Header
    assert ws.cell(1, 1).value == "상품명"
    assert ws.cell(1, 2).value == "가격"
    # Data
    assert ws.cell(2, 1).value == "상품A"
    assert ws.cell(2, 2).value == 10000
    assert ws.cell(3, 1).value == "상품B"
    assert ws.max_row == 3  # header + 2 rows


def test_export_header_bold(tmp_path):
    data = [{"col1": "val"}]
    filepath = str(tmp_path / "test.xlsx")
    export_to_excel(data, filepath)
    wb = load_workbook(filepath)
    ws = wb.active
    assert ws.cell(1, 1).font.bold is True


def test_export_auto_filter(tmp_path):
    data = [{"a": 1, "b": 2}]
    filepath = str(tmp_path / "test.xlsx")
    export_to_excel(data, filepath)
    wb = load_workbook(filepath)
    ws = wb.active
    assert ws.auto_filter.ref is not None


def test_export_empty_data(tmp_path):
    filepath = str(tmp_path / "empty.xlsx")
    result = export_to_excel([], filepath)
    assert result is None  # No file created for empty data


def test_export_korean_encoding(tmp_path):
    data = [{"이름": "홍길동", "주소": "서울시 강남구"}]
    filepath = str(tmp_path / "korean.xlsx")
    export_to_excel(data, filepath)
    wb = load_workbook(filepath)
    ws = wb.active
    assert ws.cell(2, 1).value == "홍길동"
