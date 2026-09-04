"""Phase 2 E2E 통합 테스트: API 수집 + 인증 토큰 관리"""
import json
import os
import sys
import time

sys.path.insert(0, os.path.dirname(__file__))
from utils import (
    RateLimiter, setup_logger, sanitize_filename,
    save_auth_token, load_auth_token, save_cookies, load_cookies,
)
from export_excel import export_to_excel
from scrapling.fetchers import Fetcher
from datetime import datetime


def test_e2e_api_collection(tmp_path):
    """jsonplaceholder.typicode.com API에서 posts 제목, 작성자ID를 수집하여 엑셀로 출력."""
    logger = setup_logger("e2e_phase2_api")
    # 산출물은 tmp_path 아래로. 실제 output/ 에 쓰면 도메인별 폴더 규약을 어기고
    # pytest 를 돌릴 때마다 output/ 루트에 파편이 쌓인다.
    output_dir = str(tmp_path)

    # === Step 1: API 수집 (Fetcher + resp.json()) ===
    logger.info("=== Phase 2 E2E: API Collection (jsonplaceholder) ===")
    limiter = RateLimiter(delay=0.5)
    results = []
    fetcher = Fetcher()
    page_num = 1
    limit = 10
    max_pages = 3  # 테스트용 3페이지만

    while page_num <= max_pages:
        limiter.wait()
        url = f"https://jsonplaceholder.typicode.com/posts?_page={page_num}&_limit={limit}"
        logger.info(f"API call page {page_num}: {url}")

        resp = fetcher.get(url)
        if resp.status != 200:
            logger.warning(f"Status {resp.status}, stopping")
            break

        data = resp.json()
        if not data:
            break

        for item in data:
            results.append({
                "ID": item.get("id"),
                "제목": item.get("title"),
                "작성자ID": item.get("userId"),
                "본문": item.get("body", "")[:50],  # 본문은 50자로 자름
            })

        logger.info(f"Page {page_num}: {len(data)} items, total {len(results)}")

        if len(data) < limit:
            break
        page_num += 1

    # === Step 2: 원시 데이터 저장 ===
    raw_path = os.path.join(output_dir, "raw_data_api.json")
    with open(raw_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    logger.info(f"Raw data saved: {raw_path} ({len(results)} items)")

    # === Step 3: 데이터 검증 ===
    assert len(results) == 30, f"Expected 30 items (3 pages x 10), got {len(results)}"
    empty_titles = sum(1 for r in results if not r["제목"])
    assert empty_titles == 0, f"{empty_titles} empty titles"

    # ID 순서 확인
    ids = [r["ID"] for r in results]
    assert ids == list(range(1, 31)), f"ID sequence mismatch: {ids[:5]}..."

    logger.info(f"First: {results[0]}")
    logger.info(f"Last: {results[-1]}")

    # === Step 4: 엑셀 출력 ===
    domain = sanitize_filename("jsonplaceholder_typicode_com")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f"crawl_{domain}_{timestamp}.xlsx")
    result = export_to_excel(results, excel_path)
    assert result is not None
    assert os.path.exists(excel_path)

    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    assert ws.cell(1, 1).value == "ID"
    assert ws.cell(1, 2).value == "제목"
    assert ws.max_row == 31  # header + 30 rows
    logger.info(f"Excel saved: {excel_path} ({ws.max_row - 1} rows)")

    logger.info("=== API Collection Test PASSED ===")


def test_auth_token_lifecycle(tmp_path):
    """인증 토큰 저장/로드/만료 생명주기 테스트."""
    logger = setup_logger("e2e_phase2_auth")
    # 산출물은 tmp_path 아래로. 실제 output/ 에 쓰면 도메인별 폴더 규약을 어기고
    # pytest 를 돌릴 때마다 output/ 루트에 파편이 쌓인다.
    output_dir = str(tmp_path)

    logger.info("=== Phase 2 E2E: Auth Token Lifecycle ===")

    # JWT 토큰 저장/로드
    token = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.e2e_test.sig"
    token_path = os.path.join(output_dir, "token_test.json")
    save_auth_token(token, token_path, token_type="bearer")
    loaded = load_auth_token(token_path)
    assert loaded is not None
    assert loaded["type"] == "bearer"
    assert loaded["token"] == token
    logger.info(f"Token save/load: OK")

    # 쿠키 저장/로드
    cookies = {"sessionid": "abc123", "csrftoken": "xyz789"}
    cookie_path = os.path.join(output_dir, "cookies_test.json")
    save_cookies(cookies, cookie_path)
    loaded_cookies = load_cookies(cookie_path)
    assert loaded_cookies == cookies
    logger.info(f"Cookie save/load: OK")

    # 만료 테스트 (토큰)
    old_time = time.time() - (25 * 3600)
    os.utime(token_path, (old_time, old_time))
    expired = load_auth_token(token_path)
    assert expired is None
    logger.info(f"Token expiry (24h): OK")

    # 만료 테스트 (쿠키)
    os.utime(cookie_path, (old_time, old_time))
    expired_cookies = load_cookies(cookie_path)
    assert expired_cookies is None
    logger.info(f"Cookie expiry (24h): OK")

    # 클린업
    os.remove(token_path)
    os.remove(cookie_path)

    logger.info("=== Auth Token Lifecycle Test PASSED ===")


if __name__ == "__main__":
    test_e2e_api_collection()
    print()
    test_auth_token_lifecycle()
    print("\n=== ALL PHASE 2 E2E TESTS PASSED ===")
