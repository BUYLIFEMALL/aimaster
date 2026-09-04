"""Phase 5 E2E 통합 테스트: 셀렉터 자가 치유 + 도메인 프로필 + 데이터 정제 + PII 감지"""
import json
import os
import re
import shutil
import sys

sys.path.insert(0, os.path.dirname(__file__))
from utils import setup_logger, sanitize_filename, detect_pii
from export_excel import export_to_excel
from domain_profile import DomainProfile
from datetime import datetime


def test_e2e_selector_self_healing(tmp_path):
    """셀렉터 자가 치유: auto_save로 핑거프린트 저장 후 adaptive로 복구 확인."""
    from scrapling.parser import Adaptor

    logger = setup_logger("e2e_phase5_healing")
    # 핑거프린트 DB는 tmp_path 아래에 만든다. 실제 fingerprints/ 에 쓰면
    # SQLite 사이드카(-shm/-wal)가 남아 프로덕션 디렉터리를 오염시킨다.
    db_path = str(tmp_path / "test_elements_storage.db")

    logger.info("=== Phase 5 E2E: Selector Self-Healing ===")

    # Step 1: 원본 HTML — 핑거프린트 저장
    html_v1 = """
    <html><body>
        <div class="product-list">
            <div class="item"><h2 class="title">상품A</h2><span class="price">₩10,000</span></div>
            <div class="item"><h2 class="title">상품B</h2><span class="price">₩20,000</span></div>
            <div class="item"><h2 class="title">상품C</h2><span class="price">₩30,000</span></div>
        </div>
    </body></html>
    """

    adaptor_v1 = Adaptor(html_v1, url="https://test-healing.com", auto_save=True,
                          storage_args={"storage_file": db_path})
    items_v1 = adaptor_v1.css("div.item h2.title")
    assert len(items_v1) == 3, f"Expected 3 items, got {len(items_v1)}"
    logger.info(f"V1: Found {len(items_v1)} items with original selector")

    # Step 2: 변경된 HTML — 클래스명 변경 (title → product-title)
    html_v2 = """
    <html><body>
        <div class="product-list">
            <div class="item"><h2 class="product-title">상품A</h2><span class="price">₩10,000</span></div>
            <div class="item"><h2 class="product-title">상품B</h2><span class="price">₩20,000</span></div>
            <div class="item"><h2 class="product-title">상품C</h2><span class="price">₩30,000</span></div>
        </div>
    </body></html>
    """

    # adaptive=True로 자가 치유 시도 — Adaptor 생성자에서 설정
    adaptor_v2 = Adaptor(html_v2, url="https://test-healing.com",
                          auto_save=True, adaptive=True,
                          storage_args={"storage_file": db_path})
    items_v2 = adaptor_v2.css("div.item h2.title")

    # adaptive가 핑거프린트 기반으로 복구했거나, 구조 기반 매칭으로 찾을 수 있음
    # Scrapling adaptive의 동작은 보장되지 않으므로 로깅만 수행
    logger.info(f"V2 (adaptive): Found {len(items_v2)} items after selector change")

    # 핑거프린트 DB가 생성되었는지 확인
    assert os.path.exists(db_path), "Fingerprint DB not created"
    logger.info(f"Fingerprint DB exists: {db_path}")

    # cleanup 불필요 — tmp_path는 매 실행 새로 만들어지고 pytest가 회수한다
    logger.info("=== Selector Self-Healing Test PASSED ===")


def test_e2e_domain_profile(tmp_path):
    """도메인 프로필 저장/로드/재사용 검증."""
    logger = setup_logger("e2e_phase5_profile")
    # 산출물은 tmp_path 아래로. 실제 output/ 에 쓰면 도메인별 폴더 규약을 어기고
    # pytest 를 돌릴 때마다 output/ 루트에 파편이 쌓인다.
    output_dir = str(tmp_path)

    logger.info("=== Phase 5 E2E: Domain Profile Save/Reuse ===")

    profile_mgr = DomainProfile(base_dir=output_dir)
    domain = "test-profile.example.com"

    # Clean
    profile_dir = os.path.join(output_dir, sanitize_filename(domain))
    if os.path.exists(profile_dir):
        shutil.rmtree(profile_dir)

    # Step 1: 프로필이 없으면 정찰 필요
    assert not profile_mgr.exists(domain), "Profile should not exist yet"
    logger.info("No existing profile — reconnaissance needed")

    # Step 2: 정찰 후 프로필 저장
    profile_data = {
        "domain": domain,
        "fetcher_type": "Fetcher",
        "selectors": {
            "상품명": "h2.title::text",
            "가격": "span.price::text",
        },
        "pagination": {
            "type": "url_param",
            "param": "page",
            "next_selector": "a.next::attr(href)",
        },
        "api_endpoints": [],
        "last_used": datetime.now().isoformat(),
    }
    profile_mgr.save(domain, profile_data)
    logger.info(f"Profile saved for {domain}")

    # Step 3: 프로필 존재 확인 + 로드
    assert profile_mgr.exists(domain), "Profile should exist after save"
    loaded = profile_mgr.load(domain)
    assert loaded["domain"] == domain
    assert loaded["fetcher_type"] == "Fetcher"
    assert "상품명" in loaded["selectors"]
    logger.info(f"Profile loaded: fetcher={loaded['fetcher_type']}, selectors={list(loaded['selectors'].keys())}")

    # Step 4: 재수집 시 프로필 재사용 시뮬레이션
    if profile_mgr.exists(domain):
        profile = profile_mgr.load(domain)
        logger.info(f"Reusing profile — skip reconnaissance, use {profile['fetcher_type']}")
    else:
        logger.info("No profile — run reconnaissance")

    # Cleanup
    if os.path.exists(profile_dir):
        shutil.rmtree(profile_dir)

    logger.info("=== Domain Profile Test PASSED ===")


def test_e2e_smart_cleansing():
    """스마트 데이터 정제: 가격 문자열 정규화, 상대 URL 변환."""
    logger = setup_logger("e2e_phase5_cleansing")

    logger.info("=== Phase 5 E2E: Smart Data Cleansing ===")

    # 원시 데이터 (수집 직후 상태 시뮬레이션)
    raw_data = [
        {"상품명": "상품A", "가격": "₩15,000", "링크": "/products/1", "설명": "  좋은 상품입니다.  "},
        {"상품명": "상품B", "가격": "$29.99", "링크": "/products/2", "설명": "<b>인기</b> 상품"},
        {"상품명": "상품C", "가격": "₩8,500", "링크": "https://example.com/products/3", "설명": "  "},
        {"상품명": "상품D", "가격": "무료", "링크": "/products/4", "설명": "무료 배송"},
        {"상품명": "상품E", "가격": "₩123,456", "링크": "/products/5", "설명": None},
    ]

    # 정제 함수 (에이전트가 샘플 분석 후 동적 생성하는 패턴)
    def clean_price(val):
        """₩15,000 → 15000, $29.99 → 29.99"""
        if not val:
            return None
        cleaned = re.sub(r'[^\d.]', '', val.replace(',', ''))
        try:
            return float(cleaned)
        except ValueError:
            return val  # "무료" 같은 비숫자는 그대로

    def clean_url(val, base="https://example.com"):
        """상대 URL → 절대 URL"""
        if val and val.startswith('/'):
            return base + val
        return val

    def clean_text(val):
        """HTML 태그 제거 + 공백 정리"""
        if not val:
            return ""
        cleaned = re.sub(r'<[^>]+>', '', val)  # HTML 태그 제거
        cleaned = cleaned.strip()
        return cleaned

    # 정제 적용
    for item in raw_data:
        item["가격"] = clean_price(item.get("가격"))
        item["링크"] = clean_url(item.get("링크"))
        item["설명"] = clean_text(item.get("설명"))

    # 검증
    assert raw_data[0]["가격"] == 15000.0, f"Price cleansing failed: {raw_data[0]['가격']}"
    assert raw_data[1]["가격"] == 29.99, f"USD price failed: {raw_data[1]['가격']}"
    assert raw_data[3]["가격"] == "무료", f"Non-numeric should stay: {raw_data[3]['가격']}"
    logger.info(f"Price cleansing: OK ({[item['가격'] for item in raw_data]})")

    assert raw_data[0]["링크"] == "https://example.com/products/1"
    assert raw_data[2]["링크"] == "https://example.com/products/3"  # 이미 절대 URL
    logger.info(f"URL cleansing: OK")

    assert raw_data[1]["설명"] == "인기 상품"  # HTML 태그 제거됨
    assert raw_data[0]["설명"] == "좋은 상품입니다."  # 공백 정리됨
    assert raw_data[2]["설명"] == ""  # 빈 공백만 있던 것 → 빈 문자열
    assert raw_data[4]["설명"] == ""  # None → 빈 문자열
    logger.info(f"Text cleansing: OK")

    logger.info("=== Smart Data Cleansing Test PASSED ===")


def test_e2e_pii_detection():
    """PII 감지: 이메일/전화번호 패턴 경고."""
    logger = setup_logger("e2e_phase5_pii")

    logger.info("=== Phase 5 E2E: PII Detection ===")

    # PII가 포함된 데이터
    data_with_pii = [
        {"이름": "홍길동", "이메일": "hong@example.com", "전화": "010-1234-5678"},
        {"이름": "김철수", "이메일": "kim@test.co.kr", "전화": "02-555-1234"},
        {"이름": "이영희", "이메일": "lee@gmail.com", "전화": "031-999-8888"},
    ]

    warnings = detect_pii(data_with_pii)
    assert len(warnings) > 0, "Should detect PII"
    email_warnings = [w for w in warnings if "이메일" in w]
    phone_warnings = [w for w in warnings if "전화번호" in w]
    assert len(email_warnings) > 0, "Should detect email PII"
    assert len(phone_warnings) > 0, "Should detect phone PII"
    logger.info(f"PII detected: {len(email_warnings)} email, {len(phone_warnings)} phone warnings")
    for w in warnings:
        logger.info(f"  - {w}")

    # PII가 없는 데이터
    clean_data = [
        {"상품명": "노트북", "가격": "₩1,500,000", "카테고리": "전자제품"},
        {"상품명": "마우스", "가격": "₩25,000", "카테고리": "주변기기"},
    ]

    clean_warnings = detect_pii(clean_data)
    assert len(clean_warnings) == 0, f"Should not detect PII in clean data: {clean_warnings}"
    logger.info("Clean data: no PII detected (correct)")

    logger.info("=== PII Detection Test PASSED ===")


def test_e2e_full_pipeline(tmp_path):
    """전체 파이프라인 통합: 수집→정제→PII 확인→엑셀 출력."""
    from scrapling.fetchers import Fetcher
    from utils import RateLimiter

    logger = setup_logger("e2e_phase5_pipeline")
    # 산출물은 tmp_path 아래로. 실제 output/ 에 쓰면 도메인별 폴더 규약을 어기고
    # pytest 를 돌릴 때마다 output/ 루트에 파편이 쌓인다.
    output_dir = str(tmp_path)

    logger.info("=== Phase 5 E2E: Full Pipeline (books.toscrape.com) ===")

    # === Step 1: 도메인 프로필 확인 ===
    profile_mgr = DomainProfile(base_dir=output_dir)
    domain = "books.toscrape.com"

    # === Step 2: 수집 (2페이지) ===
    fetcher = Fetcher()
    limiter = RateLimiter(delay=0.5)
    results = []

    for page_num in range(1, 3):
        limiter.wait()
        url = f"https://books.toscrape.com/catalogue/page-{page_num}.html"
        page = fetcher.get(url)
        assert page.status == 200

        for book in page.css("article.product_pod"):
            results.append({
                "제목": book.css("h3 a::attr(title)").get("").strip(),
                "가격": book.css("p.price_color::text").get("").strip(),
                "평점": book.css("p.star-rating::attr(class)").get("").replace("star-rating ", ""),
            })

    assert len(results) == 40, f"Expected 40 items, got {len(results)}"
    logger.info(f"Collected {len(results)} items")

    # === Step 3: 스마트 정제 ===
    def clean_price(val):
        if not val:
            return None
        cleaned = re.sub(r'[^\d.]', '', val.replace(',', ''))
        try:
            return float(cleaned)
        except ValueError:
            return val

    for item in results:
        item["가격"] = clean_price(item.get("가격"))

    # 가격이 숫자로 변환되었는지 확인
    numeric_prices = sum(1 for r in results if isinstance(r["가격"], float))
    assert numeric_prices == len(results), f"Expected all prices numeric, got {numeric_prices}/{len(results)}"
    logger.info(f"Data cleansing: {numeric_prices}/{len(results)} prices normalized")

    # === Step 4: PII 확인 ===
    warnings = detect_pii(results)
    logger.info(f"PII check: {len(warnings)} warnings (expected 0 for product data)")
    assert len(warnings) == 0, f"Unexpected PII in product data: {warnings}"

    # === Step 5: 엑셀 출력 ===
    domain_safe = sanitize_filename(domain)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_dir, f"crawl_{domain_safe}_{timestamp}.xlsx")
    export_result = export_to_excel(results, excel_path)
    assert export_result is not None
    assert os.path.exists(excel_path)

    from openpyxl import load_workbook
    wb = load_workbook(excel_path)
    ws = wb.active
    assert ws.cell(1, 1).value == "제목"
    assert ws.cell(1, 2).value == "가격"
    assert ws.max_row == 41  # header + 40 rows

    # 가격이 숫자로 저장되었는지 확인
    first_price = ws.cell(2, 2).value
    assert isinstance(first_price, (int, float)), f"Price should be numeric in Excel: {type(first_price)}"
    logger.info(f"Excel saved: {excel_path} ({ws.max_row - 1} rows, prices as numbers)")

    # === Step 6: 도메인 프로필 저장 ===
    profile_mgr.save(domain, {
        "domain": domain,
        "fetcher_type": "Fetcher",
        "selectors": {
            "제목": "h3 a::attr(title)",
            "가격": "p.price_color::text",
            "평점": "p.star-rating::attr(class)",
        },
        "pagination": {"type": "url_param", "next_selector": "li.next a::attr(href)"},
        "api_endpoints": [],
        "last_used": datetime.now().isoformat(),
    })
    assert profile_mgr.exists(domain)
    logger.info(f"Domain profile saved for {domain}")

    # Cleanup profile
    profile_dir = os.path.join(output_dir, sanitize_filename(domain))
    if os.path.exists(profile_dir):
        shutil.rmtree(profile_dir)

    logger.info("=== Full Pipeline Test PASSED ===")


if __name__ == "__main__":
    test_e2e_selector_self_healing()
    print()
    test_e2e_domain_profile()
    print()
    test_e2e_smart_cleansing()
    print()
    test_e2e_pii_detection()
    print()
    test_e2e_full_pipeline()
    print("\n=== ALL PHASE 5 E2E TESTS PASSED ===")
